#!/usr/bin/env python3
"""
build_category_db.py — Stage 1 (normalisation) — multi-pays.

Transforme les snapshots Nielsen (1 Excel hétérogène par pays) en UNE base
canonique au format LONG. Robuste à :
  * en-tête sur ligne variable (détectée via cellule A == 'Markets') ;
  * bloc variables détecté via motif de date dans l'en-tête (FR 'finissant le',
    UK 'w/e') -> indépendant de la langue ;
  * colonnes hiérarchiques en nombre/ordre/nom DIFFERENTS d'un onglet à l'autre
    -> on mappe par LIBELLE (synonymes), jamais par position ;
  * dernière colonne EAN variable (M / O / P / Q...) ;
  * jeton MDD propre au pays ('MDD', 'PRIVATE LABEL'...) ;
  * détection EMPIRIQUE du détail MDD (FR: non détaillée ; UK: détaillée).
"""
import datetime, re, sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DATE = re.compile(r"\d{2}/\d{2}/\d{2}")

# --- Synonymes de RÔLE (libellé de colonne -> rôle canonique), tous pays ------
ROLE_SYNONYMS = {
    "FAMILLES": "family", "TOTAL CATEGORY": "family",
    "GLOBAL PRODUCT - TYPE": "category",   # NB: position variable selon onglet
    "MARQUE PROPRIETAIRE": "manufacturer", "TRADING COMPANY": "manufacturer",
    "MARQUE": "brand", "BRAND": "brand",
    "SEGMENT": "segment",
    "DESCRIPTION PRODUIT": "flavor", "FLAVOUR OF CAKE DESSERT": "flavor",
    "WEIGHT VOLUME ACTUAL": "weight", "WEIGHT/VOLUME": "weight",
    "NUMBER IN MULTIPACK ACTUAL": "multipack", "BASE NUMBER IN MULTIPACK": "multipack",
    "BIO / NON BIO": "bio", "ORGANIC VS STANDARD": "bio", "BIO": "bio",
    "UPC": "ean",
    # Allemagne
    "CATEGORY": "family", "WARENGRUPPE": "category",
    "HERSTELLER": "manufacturer", "MARKE": "brand",
    "GESCHMACK": "flavor", "INHALT": "weight",
    # Espagne
    "CATEGORIA": "family", "FAMILIA": "category", "FABRICANTES": "manufacturer",
    "MARCA": "brand", "VARIEDAD": "flavor", "PESO": "weight",
    "UNIDADES": "multipack", "RECLAMO BIO": "bio",
    # Italie
    "COMPARTO_ECR2": "family", "FAMIGLIA_ECR3": "category",
    "FABBRICANTE": "manufacturer", "GUSTO": "flavor", "SEGMENTO": "segment",
    "FORMATO": "weight", "N PEZZI": "multipack",
}
# ordre de profondeur canonique (pour nommer le niveau independamment de la position)
ROLE_DEPTH = ["family", "category", "manufacturer", "brand", "segment",
              "multipack", "weight", "bio", "flavor", "ean"]

# --- Harmonisation parfums : libellé local (toute langue) -> parfum de BASE ----
# On scanne les jetons du libellé et on retient le parfum du PREMIER jeton reconnu
# (= parfum principal). Les composés retombent donc sur leur parfum dominant.
# (stem en MAJUSCULES sans accent ; on teste token.startswith(stem))
FLAVOR_STEMS = [
    ("ERDBEER", "Strawberry"), ("FRAISE", "Strawberry"), ("FRAGOL", "Strawberry"),
    ("FRESA", "Strawberry"), ("STRAWBERR", "Strawberry"),
    ("HIMBEER", "Raspberry"), ("FRAMBOIS", "Raspberry"), ("LAMPON", "Raspberry"),
    ("FRAMBUES", "Raspberry"), ("RASPBERR", "Raspberry"),
    ("HEIDELBEER", "Blueberry"), ("BLAUBEER", "Blueberry"), ("MYRTILL", "Blueberry"),
    ("MIRTILL", "Blueberry"), ("ARANDAN", "Blueberry"), ("BLUEBERR", "Blueberry"),
    ("BROMBEER", "Blackberry"), ("BLACKBERR", "Blackberry"), ("MURE", "Blackberry"),
    ("MUR", "Blackberry"), ("MORA", "Blackberry"), ("MORE", "Blackberry"),
    ("KIRSCH", "Cherry"), ("CERISE", "Cherry"), ("CILIEG", "Cherry"),
    ("CEREZA", "Cherry"), ("CHERR", "Cherry"),
    ("CRANBERR", "Cranberry"), ("KRANBEER", "Cranberry"),
    ("JOHANNISB", "Currant"), ("CASSIS", "Currant"), ("GROSEILL", "Currant"),
    ("RIBES", "Currant"), ("BLACKCURRANT", "Currant"), ("CURRANT", "Currant"),
    ("VANILL", "Vanilla"), ("VANIGL", "Vanilla"), ("VAINILL", "Vanilla"),
    ("BOURBON", "Vanilla"), ("VANILLA", "Vanilla"),
    ("ZITRON", "Lemon"), ("CITRON", "Lemon"), ("LIMON", "Lemon"), ("LEMON", "Lemon"),
    ("CITRIC", "Lemon"), ("CITRUS", "Lemon"),
    ("LIMETT", "Lime"), ("LIME", "Lime"),
    ("KOKOS", "Coconut"), ("COCO", "Coconut"), ("COCCO", "Coconut"), ("COCONUT", "Coconut"),
    ("PFIRSICH", "Peach"), ("PECHE", "Peach"), ("PESCA", "Peach"),
    ("MELOCOT", "Peach"), ("PEACH", "Peach"),
    ("APRIKOS", "Apricot"), ("ABRICOT", "Apricot"), ("ALBICOCC", "Apricot"),
    ("ALBARICOQ", "Apricot"), ("APRICOT", "Apricot"),
    ("MANGO", "Mango"), ("MANGUE", "Mango"),
    ("ANANAS", "Pineapple"), ("PINEAPPLE", "Pineapple"), ("PINA", "Pineapple"),
    ("MARACUJA", "Passion Fruit"), ("PASSION", "Passion Fruit"), ("PASION", "Passion Fruit"),
    ("BLUTORANGE", "Orange"), ("BLUTOR", "Orange"), ("ORANGE", "Orange"),
    ("ARANCI", "Orange"), ("NARANJA", "Orange"),
    ("MANDARIN", "Mandarin"), ("CLEMENTIN", "Mandarin"),
    ("BANAN", "Banana"), ("PLATAN", "Banana"),
    ("BRATAPFEL", "Apple"), ("GRUENAPFEL", "Apple"), ("APFEL", "Apple"), ("POMME", "Apple"),
    ("MELA", "Apple"), ("MANZAN", "Apple"), ("APPLE", "Apple"),
    ("BIRNE", "Pear"), ("POIRE", "Pear"), ("PERA", "Pear"), ("PEAR", "Pear"),
    ("FEIGE", "Fig"), ("FIGUE", "Fig"), ("FICO", "Fig"), ("FICH", "Fig"),
    ("HIGO", "Fig"), ("FIG", "Fig"),
    ("PFLAUME", "Plum"), ("CIRUELA", "Plum"), ("PRUGNA", "Plum"), ("PLUM", "Plum"),
    ("MARRON", "Chestnut"), ("MARONI", "Chestnut"), ("KASTANIE", "Chestnut"),
    ("CASTAN", "Chestnut"), ("CHESTNUT", "Chestnut"),
    ("RHABARBER", "Rhubarb"), ("RHUBARB", "Rhubarb"), ("RUIBARBO", "Rhubarb"),
    ("RABARBAR", "Rhubarb"),
    ("GINGER", "Ginger"), ("INGWER", "Ginger"), ("JENGIBRE", "Ginger"), ("ZENZERO", "Ginger"),
    ("TZATZIK", "Tzatziki"), ("ZAZIK", "Tzatziki"),
    ("DATTEL", "Date"), ("DATIL", "Date"), ("DATTER", "Date"),
    ("GRANATAPFEL", "Pomegranate"), ("GRENADE", "Pomegranate"), ("MELOGRAN", "Pomegranate"),
    ("GRANAD", "Pomegranate"), ("POMEGRAN", "Pomegranate"),
    ("KIWI", "Kiwi"),
    ("TRAUBE", "Grape"), ("RAISIN", "Grape"), ("UVA", "Grape"), ("GRAPE", "Grape"),
    ("LITSCHI", "Lychee"), ("LYCHEE", "Lychee"), ("LICHI", "Lychee"),
    ("STRACCIATELL", "Stracciatella"),
    ("SCHOKO", "Chocolate"), ("CHOCOLAT", "Chocolate"), ("CHOCO", "Chocolate"),
    ("CIOCCOLAT", "Chocolate"), ("CHOCOLATE", "Chocolate"), ("CACAO", "Chocolate"),
    ("KAKAO", "Chocolate"),
    ("KAFFEE", "Coffee"), ("CAFE", "Coffee"), ("CAFFE", "Coffee"), ("COFFEE", "Coffee"),
    ("KARAMELL", "Caramel"), ("CARAMEL", "Caramel"), ("TOFFEE", "Caramel"),
    ("HONIG", "Honey"), ("MIEL", "Honey"), ("HONEY", "Honey"),
    ("WALNUSS", "Walnut"), ("NOIX", "Walnut"), ("NOCE", "Walnut"), ("NOCI", "Walnut"),
    ("NUEZ", "Walnut"), ("NUECES", "Walnut"), ("WALNUT", "Walnut"),
    ("HASELN", "Hazelnut"), ("NOISETT", "Hazelnut"), ("NOCCIOL", "Hazelnut"),
    ("AVELLAN", "Hazelnut"), ("HAZELNUT", "Hazelnut"),
    ("PISTAZ", "Pistachio"), ("PISTACH", "Pistachio"), ("PISTACC", "Pistachio"),
    ("MANDEL", "Almond"), ("AMANDE", "Almond"), ("MANDORL", "Almond"),
    ("ALMENDR", "Almond"), ("ALMOND", "Almond"),
    ("MUESLI", "Muesli"), ("MUSLI", "Muesli"), ("CEREAL", "Muesli"), ("GRANOLA", "Muesli"),
    ("BIRCHER", "Muesli"), ("HAFER", "Muesli"), ("AVENA", "Muesli"),
    ("TIRAMISU", "Tiramisu"),
    ("CHEESECAKE", "Cheesecake"), ("KASEKUCHEN", "Cheesecake"),
    ("SACHER", "Sachertorte"),
    ("BEERENMIX", "Mixed Berries"), ("BEEREN", "Mixed Berries"), ("WALDFRUCHT", "Mixed Berries"),
    ("FRUTTI", "Mixed Berries"), ("FRUTOS", "Mixed Berries"),
    ("BOSCO", "Mixed Berries"),
    ("MACEDONIA", "Mixed Fruit"), ("MULTIFRU", "Mixed Fruit"), ("MISCHUNG", "Mixed Fruit"),
    ("FRUCHT", "Mixed Fruit"), ("FRUTTA", "Mixed Fruit"), ("FRUTAS", "Mixed Fruit"),
    ("MIXED", "Mixed Fruit"), ("MIX", "Mixed Fruit"), ("ASSORT", "Mixed Fruit"),
]
PLAIN_STEMS = ("NATUR", "PLAIN", "CLASSIC", "CLASSICO", "GRECO", "GRIEG", "GREEK",
               "GREGO", "CREMA", "CREME", "CREMOS", "CREMIG", "CREAM", "PANNA",
               "SAHNE", "STD", "PUR", "STICHFEST", "MILD", "ORIGINAL")
UNSWEET = ("NO AZUCAR", "SIN AZUCAR", "UNSWEETEN", "OHNE ZUCKER", "SENZA ZUCCHERO",
           "SANS SUCRE", "0 AZUCAR")
SWEET = ("AZUCAR", "SWEETEN", "ZUCKER", "ZUCCHER", "SUCRE", "SUGAR")


def _strip_accents(s):
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def base_flavor(raw):
    """Local flavor label (any language) -> canonical base flavor (primary note)."""
    s = _strip_accents(str(raw)).upper()
    s = re.sub(r"[^A-Z0-9]+", " ", s).strip()
    if not s:
        return None
    tokens = s.split()
    for tok in tokens:                              # primary = first recognised token
        for stem, fl in FLAVOR_STEMS:
            if tok.startswith(stem):
                return fl
    if any(p in s for p in UNSWEET):                # plain, no fruit, kept distinct
        return "Unsweetened Plain"
    if any(p in s for p in SWEET):
        return "Sweetened Plain"
    if any(tok.startswith(p) for tok in tokens for p in PLAIN_STEMS):
        return "Plain"
    return None                                     # unresolved (caller falls back)


VARIABLE_MAP = {
    # France (valeurs déjà en EUR)
    "Ventes Valeur": ("sales_value", "value", "additive"),
    "% Evol. Ventes Valeur vs An-1": ("sales_value_yoy", "value", "node_only"),
    "PDM Valeur - Product": ("value_share", "value", "derived"),
    "Ecart Abs. PDM Valeur vs An-1 - Product": ("value_share_chg", "value", "node_only"),
    "Ventes Volume EQ": ("sales_vol_eq", "volume_eq", "additive"),
    "% Evol. Ventes Volume EQ vs An-1": ("sales_vol_eq_yoy", "volume_eq", "node_only"),
    "PDM EQ - Product": ("vol_eq_share", "volume_eq", "derived"),
    "Ecart Abs. PDM EQ vs An-1 - Product": ("vol_eq_share_chg", "volume_eq", "node_only"),
    "Ventes UC": ("sales_units", "units", "additive"),
    "% Evol. Ventes UC vs An-1": ("sales_units_yoy", "units", "node_only"),
    "PDM UC - Product": ("unit_share", "units", "derived"),
    "Ecart Abs. PDM UC vs An-1 - Product": ("unit_share_chg", "units", "node_only"),
    "Prix Moyen EQ": ("avg_price_eq", "price", "derived"),
    "% Evol. Prix Moyen EQ vs An-1": ("avg_price_eq_yoy", "price", "node_only"),
    "Prix Moyen UC": ("avg_price_unit", "price", "derived"),
    "% Evol. Prix Moyen UC vs An-1": ("avg_price_unit_yoy", "price", "node_only"),
    "DN Diff": ("dn", "distr", "node_only"),
    "Ecart Abs. DN Diff vs An-1": ("dn_chg", "distr", "node_only"),
    "DV Diff": ("dv", "distr", "node_only"),
    "Ecart Abs. DV Diff vs An-1": ("dv_chg", "distr", "node_only"),
    "CAMH pondéré Diff": ("ros_value", "rotn", "node_only"),
    "% Evol. CAMH pondéré Diff vs An-1": ("ros_value_yoy", "rotn", "node_only"),
    "VMH EQ pondérées Diff": ("ros_vol", "rotn", "node_only"),
    "% Evol. VMH EQ pondérées Diff vs An-1": ("ros_vol_yoy", "rotn", "node_only"),
    "VMH UC pondérées Diff": ("ros_units", "rotn", "node_only"),
    "% Evol. VMH UC pondérées Diff vs An-1": ("ros_units_yoy", "rotn", "node_only"),
    # United Kingdom
    "Sales Value": ("sales_value_gbp", "value", "additive"),          # devise locale
    "Sales Value (EUR)": ("sales_value", "value", "additive"),        # canonique EUR
    "Sales Value % Chg YA": ("sales_value_yoy", "value", "node_only"),
    "Share of Sales Value - Product": ("value_share", "value", "derived"),
    "Share of Sales Value Chg YA - Product": ("value_share_chg", "value", "node_only"),
    "Sales (KG)": ("sales_vol_kg", "volume_kg", "additive"),
    "Sales (KG) % Chg YA": ("sales_vol_kg_yoy", "volume_kg", "node_only"),
    "Share of Sales - Sales (KG) - Product": ("vol_kg_share", "volume_kg", "derived"),
    "Share of Sales - Sales (KG) Chg YA - Product": ("vol_kg_share_chg", "volume_kg", "node_only"),
    "Sales Units": ("sales_units", "units", "additive"),
    "Sales Units % Chg YA": ("sales_units_yoy", "units", "node_only"),
    "Share of Sales - Sales Units - Product": ("unit_share", "units", "derived"),
    "Share of Sales - Sales Units Chg YA - Product": ("unit_share_chg", "units", "node_only"),
    "Price per Sales (KG)": ("avg_price_kg_gbp", "price", "derived"),
    "Price per Sales (KG) % Chg YA": ("avg_price_kg_yoy", "price", "node_only"),
    "Price per Sales Unit": ("avg_price_unit_gbp", "price", "derived"),
    "Price per Sales Unit % Chg YA": ("avg_price_unit_yoy", "price", "node_only"),
    "Numeric Distribution - Reach": ("dn", "distr", "node_only"),
    "Numeric Distribution - Reach Chg YA": ("dn_chg", "distr", "node_only"),
    "Weighted Distribution - Reach": ("dv", "distr", "node_only"),
    "Weighted Distribution - Reach Chg YA": ("dv_chg", "distr", "node_only"),
    "ROS Value - Reach": ("ros_value_gbp", "rotn", "node_only"),
    "ROS Value - Reach % Chg YA": ("ros_value_yoy", "rotn", "node_only"),
    "ROS (KG) - Reach": ("ros_vol", "rotn", "node_only"),
    "ROS (KG) - Reach % Chg YA": ("ros_vol_yoy", "rotn", "node_only"),
    "ROS Units - Reach": ("ros_units", "rotn", "node_only"),
    "ROS Units - Reach % Chg YA": ("ros_units_yoy", "rotn", "node_only"),
    "Sales Value (EUR)": ("sales_value", "value", "additive"),
    "Price per Sales (KG) (EUR)": ("avg_price_kg", "price", "derived"),
    "Price per Sales Unit (EUR)": ("avg_price_unit", "price", "derived"),
    # Allemagne (Umsatz = EUR natif ; Absatz GESAMT = volume ; Absatz in Pack = unités)
    "Umsatz": ("sales_value", "value", "additive"),
    "%-Ver. Umsatz vs VJ": ("sales_value_yoy", "value", "node_only"),
    "Marktanteil Umsatz - Product": ("value_share", "value", "derived"),
    "Abs. Ver. Marktanteil Umsatz vs VJ - Product": ("value_share_chg", "value", "node_only"),
    "Absatz (GESAMT)": ("sales_vol_de", "volume_de", "additive"),
    "%-Ver. Absatz (GESAMT) vs VJ": ("sales_vol_de_yoy", "volume_de", "node_only"),
    "Marktanteil Absatz (GESAMT) - Product": ("vol_de_share", "volume_de", "derived"),
    "Abs. Ver. Marktanteil Absatz (GESAMT) vs VJ - Product": ("vol_de_share_chg", "volume_de", "node_only"),
    "Absatz in Pack": ("sales_units", "units", "additive"),
    "%-Ver. Absatz in Pack vs VJ": ("sales_units_yoy", "units", "node_only"),
    "Marktanteil Absatz in Pack - Product": ("unit_share", "units", "derived"),
    "Abs. Ver. Marktanteil Absatz in Pack vs VJ - Product": ("unit_share_chg", "units", "node_only"),
    "Preis pro (GESAMT)": ("avg_price_vol", "price", "derived"),
    "%-Ver. Preis pro (GESAMT) vs VJ": ("avg_price_vol_yoy", "price", "node_only"),
    "Preis pro Pack": ("avg_price_unit", "price", "derived"),
    "%-Ver. Preis pro Pack vs VJ": ("avg_price_unit_yoy", "price", "node_only"),
    "Distri num Reach": ("dn", "distr", "node_only"),
    "Abs. Ver. Distri num Reach vs VJ": ("dn_chg", "distr", "node_only"),
    "Distri gew Reach": ("dv", "distr", "node_only"),
    "Abs. Ver. Distri gew Reach vs VJ": ("dv_chg", "distr", "node_only"),
    "Prop D-Umsatz Reach": ("ros_value", "rotn", "node_only"),
    "%-Ver. Prop D-Umsatz Reach vs VJ": ("ros_value_yoy", "rotn", "node_only"),
    "Prop D-Absatz (GESAMT) Reach": ("ros_vol", "rotn", "node_only"),
    "%-Ver. Prop D-Absatz (GESAMT) Reach vs VJ": ("ros_vol_yoy", "rotn", "node_only"),
    "Prop D-Absatz in Pack Reach": ("ros_units", "rotn", "node_only"),
    "%-Ver. Prop D-Absatz in Pack Reach vs VJ": ("ros_units_yoy", "rotn", "node_only"),
    # Espagne (Vtas Valor = EUR natif ; Vtas KGS = même base que UK)
    "Vtas Valor": ("sales_value", "value", "additive"),
    "% Var Vtas Valor vs Año Ant": ("sales_value_yoy", "value", "node_only"),
    "Part. de Vtas Valor - Product": ("value_share", "value", "derived"),
    "Var Part. de Vtas Valor vs Año Ant - Product": ("value_share_chg", "value", "node_only"),
    "Vtas (KGS)": ("sales_vol_kg", "volume_kg", "additive"),
    "% Var Vtas (KGS) vs Año Ant": ("sales_vol_kg_yoy", "volume_kg", "node_only"),
    "Part. de Vtas (KGS) - Product": ("vol_kg_share", "volume_kg", "derived"),
    "Var Part. de Vtas (KGS) vs Año Ant - Product": ("vol_kg_share_chg", "volume_kg", "node_only"),
    "Vtas Unds": ("sales_units", "units", "additive"),
    "% Var Vtas Unds vs Año Ant": ("sales_units_yoy", "units", "node_only"),
    "Part. de Vtas Unds - Product": ("unit_share", "units", "derived"),
    "Var Part. de Vtas Unds vs Año Ant - Product": ("unit_share_chg", "units", "node_only"),
    "Precio (KGS) Prom.": ("avg_price_kg", "price", "derived"),
    "% Var Precio (KGS) Prom. vs Año Ant": ("avg_price_kg_yoy", "price", "node_only"),
    "Precio Unds Prom.": ("avg_price_unit", "price", "derived"),
    "% Var Precio Unds Prom. vs Año Ant": ("avg_price_unit_yoy", "price", "node_only"),
    "Dist. Num. Reach": ("dn", "distr", "node_only"),
    "Var Dist. Num. Reach vs Año Ant": ("dn_chg", "distr", "node_only"),
    "Dist. Pond. Reach": ("dv", "distr", "node_only"),
    "Var Dist. Pond. Reach vs Año Ant": ("dv_chg", "distr", "node_only"),
    "Ratio de Vtas Valor Reach": ("ros_value", "rotn", "node_only"),
    "% Var Ratio de Vtas Valor vs Año Ant": ("ros_value_yoy", "rotn", "node_only"),
    "Ratio de Vtas (KGS) Reach": ("ros_vol", "rotn", "node_only"),
    "% Var Ratio de Vtas (KGS) vs Año Ant": ("ros_vol_yoy", "rotn", "node_only"),
    "Ratio de Vtas Unds Reach": ("ros_units", "rotn", "node_only"),
    "% Var Ratio de Vtas Unds vs Año Ant": ("ros_units_yoy", "rotn", "node_only"),
    # Italie (V. Valore = EUR natif ; V.(ALL) = volume base IT ; V. Confezioni = unités)
    "V. Valore": ("sales_value", "value", "additive"),
    "Var.% V. Valore Anno prec.": ("sales_value_yoy", "value", "node_only"),
    "Quota Val. - Product": ("value_share", "value", "derived"),
    "Var.Ass. Quota Val. Anno prec. - Product": ("value_share_chg", "value", "node_only"),
    "V. (ALL)": ("sales_vol_it", "volume_it", "additive"),
    "Var.% V. (ALL) Anno prec.": ("sales_vol_it_yoy", "volume_it", "node_only"),
    "Quota in (ALL) - Product": ("vol_it_share", "volume_it", "derived"),
    "Var.Ass. Quota in (ALL) Anno prec. - Product": ("vol_it_share_chg", "volume_it", "node_only"),
    "V. Confezioni": ("sales_units", "units", "additive"),
    "Var.% V. Confezioni Anno prec.": ("sales_units_yoy", "units", "node_only"),
    "Quota Conf. - Product": ("unit_share", "units", "derived"),
    "Var.Ass. Quota Conf. Anno prec. - Product": ("unit_share_chg", "units", "node_only"),
    "Prezzo Medio (ALL)": ("avg_price_vol", "price", "derived"),
    "Var.% Prezzo Medio (ALL) Anno prec.": ("avg_price_vol_yoy", "price", "node_only"),
    "Prezzo Medio Conf.": ("avg_price_unit", "price", "derived"),
    "Var.% Prezzo Medio Conf. Anno prec.": ("avg_price_unit_yoy", "price", "node_only"),
    "Distr. Num. Reach": ("dn", "distr", "node_only"),
    "Var.Ass. Distr. Num. Reach Anno prec.": ("dn_chg", "distr", "node_only"),
    "Distr.Pond. Reach": ("dv", "distr", "node_only"),
    "Var.Ass. Distr.Pond. Reach Anno prec.": ("dv_chg", "distr", "node_only"),
    "V. Val. per pdv ponderate Reach": ("ros_value", "rotn", "node_only"),
    "Var.% V. Val. per pdv ponderate Anno prec.": ("ros_value_yoy", "rotn", "node_only"),
    "(ALL) per pdv ponderate Reach": ("ros_vol", "rotn", "node_only"),
    "Var.% V. (ALL) per pdv ponderate Anno prec.": ("ros_vol_yoy", "rotn", "node_only"),
    "V. Conf. per pdv ponderate Reach": ("ros_units", "rotn", "node_only"),
    "Var.% V. Conf. per pdv ponderate Anno prec.": ("ros_units_yoy", "rotn", "node_only"),
}

COUNTRIES = {
    "FR": {
        "file": "1781707654138_Snapshot_Olga_Yoghurts_France_Apr_26.xlsx",
        "skip": {"Menu", "Channels shares"},
        "market_map": {"Greek": "YALG", "Kefir": "Kefir", "Bifidus": "Bifidus"},
        "channel_map": {"HMSM + PROXI + DRIVE + SDMP": "Total", "ENSEIGNES HM": "Hyper"},
        "mdd_token": "MDD", "mdd_exact": False,
        "flavor_map": {"NATURE": "Plain", "VANILLE": "Vanilla", "CITRON": "Lemon",
                       "NOIX COCO": "Coconut", "LIT DE FRAISE": "Strawberry",
                       "LIT DE FRAISES": "Strawberry", "PULPE DE CITRON": "Lemon",
                       "LIT DE MYRTILLES": "Blueberry", "STRACCIATELLA": "Stracciatella",
                       "CHOCOLAT & GRANOLA": "Chocolate"},
    },
    "UK": {
        "file": "1781708349379_Snapshot_Olga_Yoghurts_United_Kingdom_Apr_26.xlsx",
        "skip": {"Menu", "Channels shares"},
        "market_map": {"Greek": "YALG", "Kefir": "Kefir", "Bifidus": "Bifidus"},
        "channel_map": {"GB_Tot Cov": "Total",
                        "GB_Groc Mults_Out of Town_Megastores": "Hyper"},
        "mdd_token": "PRIVATE LABEL",
        "flavor_map": {},  # déjà en anglais canonique (PLAIN, VANILLA, COCONUT...)
    },
    "DE": {
        "file": "1781708930453_Snapshot_Olga_Yoghurts_Germany_Apr_26.xlsx",
        "skip": {"Menu", "Channels shares"},
        "market_map": {"Greek": "YALG", "Kefir": "Kefir", "Bifidus": "Bifidus"},
        "channel_map": {"LEH+DM+C&C": "Total", "VM gross": "Hyper"},
        "mdd_token": "HANDEL",
        "flavor_map": {"NATUR": "Plain", "VANILLE": "Vanilla", "ZITRONE": "Lemon",
                       "LEMON": "Lemon", "KOKOS": "Coconut",
                       "ERDBEER": "Strawberry", "ERDBEERE": "Strawberry",
                       "HIMBEER": "Raspberry", "HIMBEERE": "Raspberry",
                       "HEIDELBEER": "Blueberry", "BLAUBEER": "Blueberry",
                       "BROMBEER": "Blackberry", "KIRSCH": "Cherry",
                       "ANANAS": "Pineapple", "MANGO": "Mango", "ORANGE": "Orange",
                       "MARACUJA": "Passion Fruit", "PFIRSICH": "Peach",
                       "MISCHUNG": "Mixed", "DIVERSE": "Various"},
    },
    "ES": {
        "file": "Snapshot_Olga_Yoghurts_Spain_Apr_26.xlsx",
        "skip": {"Menu", "Channels shares"},
        "market_map": {"Greek": "YALG", "Kefir": "Kefir", "Bifidus": "Bifidus"},
        "channel_map": {"HIPER+SUPER+INDEP+EESS+ONLINE": "Total", "HIPER": "Hyper"},
        "mdd_token": "M.DISTRI.",
        "flavor_map": {"NATURAL": "Plain", "VAINILLA": "Vanilla", "LIMON": "Lemon",
                       "FRESA": "Strawberry", "FRESAS": "Strawberry", "CEREZA": "Cherry",
                       "MANGO": "Mango", "MELOCOTON": "Peach", "MANDARINA": "Mandarin",
                       "NUECES": "Walnut", "CITRICOS": "Citrus", "MACEDONIA": "Fruit Mix",
                       "STRACCIATELLA": "Stracciatella", "AZUCARADO": "Sweetened Plain",
                       "NO.AZUCARADO": "Unsweetened Plain"},
    },
    "IT": {
        "file": "Snapshot_Olga_Yoghurts_Italy_Apr_26.xlsx",
        "skip": {"Menu", "Channels shares"},
        "market_map": {"Greek": "YALG", "Kefir": "Kefir", "Bifidus": "Bifidus"},
        "channel_map": {"Totale Italia (I+S+LIS+DIS+SD)": "Total",
                        "Hypermarkets - 1056973": "Hyper"},
        "mdd_token": "PL", "mdd_exact": True,   # 'PL' court -> match exact obligatoire
        "flavor_map": {"VANIGLIA": "Vanilla", "FRAGOLA": "Strawberry", "LIMONE": "Lemon",
                       "COCCO": "Coconut", "PESCA": "Peach", "MIRTILLO": "Blueberry",
                       "LAMPONE": "Raspberry", "NOCCIOLA": "Hazelnut", "CAFFE'": "Coffee",
                       "FRUTTI DI BOSCO": "Mixed Berries", "CREMA": "Cream",
                       "STD": "Plain", "MELA CANNELLA": "Apple Cinnamon",
                       "+MELA CANNELLA": "Apple Cinnamon", "MIELE+NOCI": "Honey Walnut",
                       "MIELE.NOCI": "Honey Walnut", "SACHERTORTE": "Sachertorte",
                       "CHOCO COCCO": "Choco Coconut"},
    },
}


def _year(v):
    m = DATE.search(v) if isinstance(v, str) else None
    if m:
        y = int(m.group(0)[-2:]); return 2000 + y
    if isinstance(v, (datetime.datetime, datetime.date)): return v.year
    return None


def detect_layout(rows):
    hdr = next(i for i, r in enumerate(rows) if r and r[0] == "Markets")
    hrow, namerow = rows[hdr], rows[hdr - 1]
    firstvar = next((j for j, v in enumerate(hrow)
                     if isinstance(v, str) and DATE.search(v)), len(hrow))
    hier = [(j, str(hrow[j]).strip()) for j in range(firstvar) if hrow[j] is not None]
    var_cols, last_name = [], None
    for j in range(firstvar, len(hrow)):
        if not (isinstance(hrow[j], str) and DATE.search(hrow[j])):
            continue
        last_name = namerow[j] if namerow[j] is not None else last_name  # propage fusion
        var_cols.append((j, str(last_name).strip(), _year(hrow[j])))
    return hdr, hier, var_cols


# mots-clés pour retrouver le fichier d'un pays quel que soit son nom exact
FILE_KEYWORDS = {
    "FR": ["france"],
    "UK": ["kingdom", "united", "_uk", "uk_"],
    "DE": ["germany", "deutschland"],
    "ES": ["spain", "espana", "españa"],
    "IT": ["italy", "italia"],
}


def find_file(root, code, cfg):
    """Trouve l'Excel d'un pays : nom exact, sinon premier .xlsx contenant un mot-clé."""
    root = Path(root)
    exact = root / cfg["file"]
    if exact.exists():
        return exact
    for p in sorted(root.glob("*.xlsx")):
        name = p.name.lower()
        if any(k in name for k in FILE_KEYWORDS.get(code, [])):
            return p
    return None


def normalize_country(code, cfg, path):
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        if sheet in cfg["skip"] or sheet not in cfg["market_map"]:
            continue
        market = cfg["market_map"][sheet]
        rows = list(wb[sheet].iter_rows(values_only=True))
        hdr, hier, var_cols = detect_layout(rows[:15])
        role_idx = {}
        for j, lbl in hier:                       # premier gagnant : gère les libellés dupliqués
            role = ROLE_SYNONYMS.get(lbl)
            if role and role not in role_idx:
                role_idx[role] = j
        idx_role = {j: ROLE_SYNONYMS.get(lbl) for j, lbl in hier}
        hier_j = [j for j, _ in hier if j != 0]

        for r in rows[hdr + 1:]:
            if not r or r[0] not in cfg["channel_map"]:
                continue
            channel = cfg["channel_map"][r[0]]
            filled = [j for j in hier_j if j < len(r) and r[j] is not None]
            if not filled:
                continue
            deepest = max(filled)
            role = idx_role.get(deepest)
            level = role or get_column_letter(deepest + 1)

            def cell(rl):
                j = role_idx.get(rl); return r[j] if (j is not None and j < len(r)) else None

            manuf = cell("manufacturer")
            up = str(manuf).upper().strip() if manuf is not None else ""
            tok = cfg["mdd_token"].upper()
            exact = cfg.get("mdd_exact", len(tok) < 4)
            is_mdd = (up == tok) if exact else (tok in up)
            raw_fl = cell("flavor")
            # base-flavor (primary note) first; curated map for the rest; else cleaned label
            flavor = (base_flavor(raw_fl)
                      or cfg["flavor_map"].get(str(raw_fl).strip().upper())
                      or str(raw_fl).strip().title()) if raw_fl is not None else None
            depth_rank = ROLE_DEPTH.index(role) if role in ROLE_DEPTH else -1
            # cross-country group: unify all private-label tokens under one label
            manuf_group = "Private Label" if is_mdd else manuf
            base = dict(country=code, channel=channel, market=market, level=level,
                        depth=depth_rank, is_mdd=is_mdd, manufacturer=manuf,
                        manufacturer_group=manuf_group,
                        brand=cell("brand"), segment=cell("segment"), flavor_en=flavor,
                        weight=cell("weight"), multipack=cell("multipack"),
                        bio=cell("bio"), ean=cell("ean"))
            for j, vlabel, yr in var_cols:
                if j >= len(r) or r[j] is None:
                    continue
                vc = VARIABLE_MAP.get(vlabel)
                if not vc:
                    continue
                base2 = dict(base); base2.update(variable=vc[0], var_family=vc[1],
                                                  agg_rule=vc[2], year=yr, value=r[j])
                out.append(base2)
    df = pd.DataFrame(out)
    # mdd_detailed EMPIRIQUE : la MDD descend-elle sous le fabricant dans cette base ?
    mddmax = (df[df.is_mdd].groupby(["country", "market"]).depth.max()
              .rename("mdd_max_depth").reset_index())
    df = df.merge(mddmax, on=["country", "market"], how="left")
    man_depth = ROLE_DEPTH.index("manufacturer")
    df["mdd_detailed"] = df.mdd_max_depth > man_depth
    return df


if __name__ == "__main__":
    root = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads"
    out = sys.argv[2] if len(sys.argv) > 2 else "category_db.parquet"
    parts = []
    for c, cfg in COUNTRIES.items():
        path = find_file(root, c, cfg)
        if not path:
            print(f"  [skip] {c} : aucun .xlsx trouvé (mots-clés: {FILE_KEYWORDS.get(c)})")
            continue
        print(f"  [ok]   {c} : {path.name}")
        parts.append(normalize_country(c, cfg, path))
    db = pd.concat(parts, ignore_index=True)
    db.to_parquet(out, index=False)
    print(f"OK -> {out}  ({len(db):,} lignes, {db.country.nunique()} pays)")
    print(db.groupby("country").size().rename("lignes").to_string())
